import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
import win32com.client  # Requires pywin32 installed (Windows only)
import webbrowser
import pandas as pd
import os
import sys
import getpass
import subprocess
import time
import shutil
import json
import fitz  # PyMuPDF
from PIL import Image, ImageTk
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import threading
import zipfile

from python.config import (
    load_config, show_config_warning, save_config,
    load_columns, save_columns, DEFAULT_CONFIG, DEFAULT_COLUMNS
)

from python.excelhandler import (
    init_excel_settings, load_excel, safe_load_excel,
    save_excel, export_excel, acquire_lock,
    release_lock, save_excel_with_lock
)

from python.pdfhandler import (
    init_pdf_settings, find_pdf, generate_pdf_thumbnail
)

from python.watchdoghandler import (
    init_watchdog_settings, load_columns_json, ExcelHandler
)

from python.multiselect import MultiSelectDropdown
# Diagram import functions
from python.diagram.createui import create_styles, stripe_rows, create_ui, update_headers, refresh_table
from python.diagram.createfilter import create_filters, clear_all_filters, apply_filters, export_filtered
from python.diagram.setting import open_settings, manage_columns
from python.diagram.add import open_add_window, save_entry
from python.diagram.contextmenu import show_context_menu, delete_selected_row, edit_selected_row, save_edited_entry, open_pdf_preview

# Load config
config = load_config()
EXCEL_PATH = config["excel_path"]
PDF_DIR = config["pdf_dir"]

# Build columns.json path
excel_dir = os.path.dirname(EXCEL_PATH)
parent_dir = os.path.dirname(excel_dir)
COLUMNS_FILE = os.path.join(parent_dir, "json/columns.json")

columns_data = load_columns(COLUMNS_FILE, DEFAULT_COLUMNS)

DEFAULT_LANG = config.get("language", "Japanese")
os.makedirs(PDF_DIR, exist_ok=True)

COLUMNS = columns_data["english"]
JAPANESE_COLUMNS = columns_data["japanese"]

# Language text
LANG_FILE = os.path.join(parent_dir, "json/lang.json")
with open(LANG_FILE, "r", encoding="utf-8") as f:
    LANG_TEXT = json.load(f)

# Dropdown options
DROPDOWN_FILE = os.path.join(parent_dir, "json/dropdowns.json")
try:
    with open(DROPDOWN_FILE, "r", encoding="utf-8") as f: 
        dropdown_options = json.load(f)
except FileNotFoundError:
    dropdown_options = {}

# Initialize Excel module with these settings
init_excel_settings(EXCEL_PATH, COLUMNS, JAPANESE_COLUMNS, LANG_TEXT, DEFAULT_LANG)

# Initialize PDF module with PDF directory
init_pdf_settings(PDF_DIR)

# Initialize watchdog module with settings
init_watchdog_settings(EXCEL_PATH, COLUMNS_FILE, LANG_TEXT, DROPDOWN_FILE, safe_load_excel)

# ===============================
# MAIN APP
# ===============================
class DiagramApp(tk.Tk):
    def __init__(self):
        super().__init__()

        # --- Language and text resources ---
        self.lang = DEFAULT_LANG
        self.text = LANG_TEXT[self.lang]

        # --- Data ---
        self.df = load_excel()
        self.columns_data = columns_data

        # --- Column visibility (default all True if not in JSON) ---
        if "visible" in self.columns_data:
            self.columns_visibility = self.columns_data["visible"]
        else:
            self.columns_visibility = {col: True for col in self.columns_data["english"]}

        # --- Dropdown options owned by app ---
        self.dropdown_options = dropdown_options

        # --- Window setup ---
        self.title(self.text["app_title"])
        self.geometry("1800x900")
        self.state("zoomed")  # start maximized

        # --- Styles and UI ---
        create_styles(self) 
        create_ui(self, self.columns_data["english"], LANG_TEXT)

        # --- Table setup ---
        update_headers(self, COLUMNS, LANG_TEXT) 
        refresh_table(self, self.df, COLUMNS, LANG_TEXT)

        # Start watchdog
        self.start_excel_watcher(EXCEL_PATH)

    def t(self, key):
        return LANG_TEXT[self.lang][key]

    # ---------- WATCHDOG ----------
    def start_excel_watcher(self, filepath):
        handler = ExcelHandler(filepath, self)
        observer = Observer()
        watch_dir = os.path.dirname(filepath) or "."
        observer.schedule(handler, path=watch_dir, recursive=False)
        observer.start()
        threading.Thread(target=lambda: observer.join(), daemon=True).start()

    # ---------- Filters ----------
    def refresh_table(self, df):
        refresh_table(self, df, COLUMNS, LANG_TEXT)

    def update_headers(self):
        update_headers(self, COLUMNS, LANG_TEXT)

    def clear_all_filters(self):
        clear_all_filters(self)

    def apply_filters(self):
        apply_filters(self, LANG_TEXT, COLUMNS)
    
    def export_filtered(self):
        export_filtered(self, LANG_TEXT)

    def open_settings(self): 
        open_settings(self, EXCEL_PATH, PDF_DIR, LANG_TEXT, COLUMNS)

    # ---------- Right-click ----------
    def show_context_menu(self, event):
        show_context_menu(self, event, LANG_TEXT, COLUMNS)

    def delete_selected_row(self):
        delete_selected_row(self, LANG_TEXT, COLUMNS)

    def edit_selected_row(self):
        edit_selected_row(self, LANG_TEXT, COLUMNS)

    def save_edited_entry(self, win, fields, pdf_var, original_search_no):
        save_edited_entry(self, win, fields, pdf_var, original_search_no,
                          EXCEL_PATH, PDF_DIR, LANG_TEXT, COLUMNS)

    # ---------- PDF Preview ----------
    def open_pdf_preview(self, event): 
        open_pdf_preview(self, event, LANG_TEXT)

    def manage_columns(self): 
        manage_columns(self, LANG_TEXT, COLUMNS, JAPANESE_COLUMNS, COLUMNS_FILE)

    # ---------- Add Entry ----------
    def open_add_window(self):
        open_add_window(self, LANG_TEXT, COLUMNS, PDF_DIR, EXCEL_PATH, DROPDOWN_FILE)

    def save_entry(self, win, fields, pdf_var):
        save_entry(self, win, fields, pdf_var, LANG_TEXT, COLUMNS, PDF_DIR, EXCEL_PATH, DROPDOWN_FILE)

# ===============================
# RUN
# ===============================
if __name__ == "__main__":
    DiagramApp().mainloop()
